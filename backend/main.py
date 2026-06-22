from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import json, io, re, os
from datetime import datetime
from db import db, init_db, CATEGORIES

RESEND_API_KEY  = os.environ.get("RESEND_API_KEY", "")
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "*").split(",")

app = FastAPI(title="MiniCRM API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup():
    init_db()


# ── Helpers ───────────────────────────────────────────────────────────────────

def customer_full(conn, customer_id: int):
    row = conn.execute("SELECT * FROM customers WHERE id = %s", (customer_id,)).fetchone()
    if not row:
        return None
    c = dict(row)
    c["kinder"]             = json.loads(c.get("kinder") or "[]")
    c["ap_scores"]          = json.loads(c.get("ap_scores") or "{}")
    c["anfrage_kategorien"] = json.loads(c.get("anfrage_kategorien") or "[]")
    raw_kq = c.get("kontaktquelle") or "[]"
    try:
        c["kontaktquelle"] = json.loads(raw_kq) if raw_kq.startswith("[") else ([raw_kq] if raw_kq else [])
    except Exception:
        c["kontaktquelle"] = []

    cats = conn.execute(
        "SELECT category, status, notizen, updated_at FROM customer_categories WHERE customer_id = %s",
        (customer_id,)
    ).fetchall()
    cat_map = {r["category"]: dict(r) for r in cats}
    for cat in CATEGORIES:
        if cat not in cat_map:
            cat_map[cat] = {"category": cat, "status": "nicht_besprochen", "notizen": "", "updated_at": None}
    c["categories"] = cat_map

    appts = conn.execute(
        "SELECT id, datum, status, notizen, created_at FROM appointments WHERE customer_id = %s ORDER BY created_at DESC",
        (customer_id,)
    ).fetchall()
    c["appointments"] = [dict(a) for a in appts]

    return c


# ── Models ────────────────────────────────────────────────────────────────────

class CustomerUpsert(BaseModel):
    p1_anrede:            Optional[str] = None
    p1_vorname:           str = ""
    p1_nachname:          str = ""
    p1_strasse:           Optional[str] = None
    p1_plz:               Optional[str] = None
    p1_ort:               Optional[str] = None
    p1_geburtsdatum:      Optional[str] = None
    p1_familienstand:     Optional[str] = None
    p1_hobby:             Optional[str] = None
    p1_haustiere:         Optional[str] = None
    p1_beruf:             Optional[str] = None
    p1_abschluss:         Optional[str] = None
    p1_gehalt_netto:      Optional[float] = None
    p1_gehalt_brutto:     Optional[float] = None
    p1_arbeitgeber:       Optional[str] = None
    p1_arbeitsverhaeltnis: Optional[str] = None
    p1_gkv_anbieter:      Optional[str] = None
    p1_gkv_stand:         Optional[str] = None
    p1_raucher:           Optional[int] = 0
    p1_email:             Optional[str] = None
    p1_handy:             Optional[str] = None
    p1_festnetz:          Optional[str] = None
    p1_iban:              Optional[str] = None
    p2_anrede:            Optional[str] = None
    p2_vorname:           Optional[str] = None
    p2_nachname:          Optional[str] = None
    p2_strasse:           Optional[str] = None
    p2_plz:               Optional[str] = None
    p2_ort:               Optional[str] = None
    p2_geburtsdatum:      Optional[str] = None
    p2_familienstand:     Optional[str] = None
    p2_hobby:             Optional[str] = None
    p2_haustiere:         Optional[str] = None
    p2_beruf:             Optional[str] = None
    p2_abschluss:         Optional[str] = None
    p2_gehalt_netto:      Optional[float] = None
    p2_gehalt_brutto:     Optional[float] = None
    p2_arbeitgeber:       Optional[str] = None
    p2_arbeitsverhaeltnis: Optional[str] = None
    p2_gkv_anbieter:      Optional[str] = None
    p2_gkv_stand:         Optional[str] = None
    p2_raucher:           Optional[int] = 0
    p2_email:             Optional[str] = None
    p2_handy:             Optional[str] = None
    p2_festnetz:          Optional[str] = None
    p2_iban:              Optional[str] = None
    kinder:               Optional[list] = []
    ap_scores:            Optional[dict] = {}
    qualitaet_absicherung: Optional[str] = None
    selbstbeteiligung:    Optional[str] = None
    budget_absicherung_monatlich: Optional[float] = None
    budget_absicherung_einmalig:  Optional[float] = None
    budget_ansparung_monatlich:   Optional[float] = None
    budget_ansparung_einmalig:    Optional[float] = None
    budget_notizen:               Optional[str] = None
    hhr_einnahmen_gehalt:    Optional[float] = None
    hhr_einnahmen_zuschuesse: Optional[float] = None
    hhr_einnahmen_weitere:   Optional[float] = None
    hhr_ausgaben_gesamt:     Optional[float] = None
    hhr_freier_betrag:       Optional[float] = None
    hhr_notizen:             Optional[str] = None
    schaeden_notizen:       Optional[str] = None
    empfehlung_notizen:     Optional[str] = None
    bewertung_google:       Optional[int] = None
    bewertung_trustpilot:   Optional[int] = None
    bewertung_facebook:     Optional[int] = None
    folgetermin_datum:      Optional[str] = None
    folgetermin_notizen:    Optional[str] = None
    berater:                Optional[str] = None
    eingruppierung:         Optional[str] = None
    notizen:                Optional[str] = None
    kontaktquelle:           Optional[list] = []
    kontaktquelle_sonstiges: Optional[str] = ""
    anfrage_kategorien:      Optional[list] = []


class CategoryUpdate(BaseModel):
    status:  str
    notizen: Optional[str] = ""


class AppointmentCreate(BaseModel):
    datum:   str
    status:  Optional[str] = ""
    notizen: Optional[str] = ""


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return {"ok": True}


@app.get("/api/termine-uebersicht")
def termine_uebersicht():
    with db() as conn:
        rows = conn.execute("""
            SELECT c.id, c.p1_vorname, c.p1_nachname, c.berater,
                   c.folgetermin_datum, c.folgetermin_notizen,
                   c.p1_handy, c.p1_email,
                   (SELECT a.status FROM appointments a
                    WHERE a.customer_id = c.id ORDER BY a.created_at DESC LIMIT 1) AS letzter_status,
                   (SELECT a.datum FROM appointments a
                    WHERE a.customer_id = c.id ORDER BY a.created_at DESC LIMIT 1) AS letzter_termin_datum
            FROM customers c
            WHERE c.folgetermin_datum IS NOT NULL AND c.folgetermin_datum != ''
            ORDER BY c.folgetermin_datum ASC
        """).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/customers")
def list_customers():
    with db() as conn:
        rows = conn.execute("""
            SELECT c.id, c.created_at, c.updated_at,
                   c.p1_vorname, c.p1_nachname, c.p1_email, c.p1_handy, c.p1_strasse, c.p1_plz, c.p1_ort,
                   c.p2_vorname, c.p2_nachname,
                   c.berater, c.folgetermin_datum, c.kontaktquelle,
                   c.bewertung_google,
                   (SELECT COUNT(*) FROM customer_categories cc
                    WHERE cc.customer_id = c.id AND cc.status = 'besprochen') AS besprochen,
                   (SELECT COUNT(*) FROM customer_categories cc
                    WHERE cc.customer_id = c.id AND cc.status = 'offen') AS offen,
                   (SELECT COUNT(*) FROM customer_categories cc
                    WHERE cc.customer_id = c.id AND cc.status = 'kein_interesse') AS kein_interesse
            FROM customers c
            ORDER BY c.updated_at DESC
        """).fetchall()
        result = []
        for r in rows:
            row = dict(r)
            raw_kq = row.get("kontaktquelle") or "[]"
            try:
                row["kontaktquelle"] = json.loads(raw_kq) if raw_kq.startswith("[") else ([raw_kq] if raw_kq else [])
            except Exception:
                row["kontaktquelle"] = []
            result.append(row)
        return result


@app.post("/api/customers", status_code=201)
def create_customer(data: CustomerUpsert):
    with db() as conn:
        d = data.model_dump()
        anfrage_kats = d.get("anfrage_kategorien") or []
        d["kinder"]             = json.dumps(d.get("kinder") or [])
        d["ap_scores"]          = json.dumps(d.get("ap_scores") or {})
        d["anfrage_kategorien"] = json.dumps(anfrage_kats)
        d["kontaktquelle"]      = json.dumps(d.get("kontaktquelle") or [])
        cols         = ", ".join(d.keys())
        placeholders = ", ".join(f"%({k})s" for k in d.keys())
        cur = conn.execute(
            f"INSERT INTO customers ({cols}) VALUES ({placeholders}) RETURNING id", d
        )
        customer_id = cur.fetchone()["id"]
        for cat in anfrage_kats:
            if cat in CATEGORIES:
                conn.execute("""
                    INSERT INTO customer_categories (customer_id, category, status, notizen, updated_at)
                    VALUES (%s, %s, 'offen', '', NOW())
                    ON CONFLICT (customer_id, category) DO UPDATE SET status='offen', updated_at=NOW()
                """, (customer_id, cat))
        return customer_full(conn, customer_id)


@app.get("/api/customers/{customer_id}")
def get_customer(customer_id: int):
    with db() as conn:
        c = customer_full(conn, customer_id)
        if not c:
            raise HTTPException(404, "Kunde nicht gefunden")
        return c


@app.put("/api/customers/{customer_id}")
def update_customer(customer_id: int, data: CustomerUpsert):
    with db() as conn:
        existing = conn.execute("SELECT id FROM customers WHERE id = %s", (customer_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Kunde nicht gefunden")
        d = data.model_dump()
        d["kinder"]             = json.dumps(d.get("kinder") or [])
        d["ap_scores"]          = json.dumps(d.get("ap_scores") or {})
        d["anfrage_kategorien"] = json.dumps(d.get("anfrage_kategorien") or [])
        d["kontaktquelle"]      = json.dumps(d.get("kontaktquelle") or [])
        set_clause = ", ".join(f"{k} = %({k})s" for k in d.keys())
        set_clause += ", updated_at = NOW()"
        conn.execute(
            f"UPDATE customers SET {set_clause} WHERE id = %(id)s",
            {**d, "id": customer_id}
        )
        return customer_full(conn, customer_id)


@app.delete("/api/customers/{customer_id}", status_code=204)
def delete_customer(customer_id: int):
    with db() as conn:
        conn.execute("DELETE FROM customers WHERE id = %s", (customer_id,))


@app.put("/api/customers/{customer_id}/categories/{category}")
def update_category(customer_id: int, category: str, data: CategoryUpdate):
    if category not in CATEGORIES:
        raise HTTPException(400, f"Unbekannte Kategorie: {category}")
    with db() as conn:
        conn.execute("""
            INSERT INTO customer_categories (customer_id, category, status, notizen, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (customer_id, category)
            DO UPDATE SET status = EXCLUDED.status, notizen = EXCLUDED.notizen, updated_at = NOW()
        """, (customer_id, category, data.status, data.notizen or ""))
        return {"ok": True}


@app.post("/api/customers/{customer_id}/appointments", status_code=201)
def add_appointment(customer_id: int, data: AppointmentCreate):
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO appointments (customer_id, datum, status, notizen) VALUES (%s, %s, %s, %s) RETURNING id",
            (customer_id, data.datum, data.status or "", data.notizen or "")
        )
        appt_id = cur.fetchone()["id"]
        return {"id": appt_id, "datum": data.datum, "status": data.status, "notizen": data.notizen}


@app.delete("/api/customers/{customer_id}/appointments/{appt_id}", status_code=204)
def delete_appointment(customer_id: int, appt_id: int):
    with db() as conn:
        conn.execute("DELETE FROM appointments WHERE id = %s AND customer_id = %s", (appt_id, customer_id))


# ── Settings ──────────────────────────────────────────────────────────────────

class SettingUpsert(BaseModel):
    value: str

@app.get("/api/settings/{key}")
def get_setting(key: str):
    with db() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = %s", (key,)).fetchone()
        if not row:
            raise HTTPException(404, f"Setting '{key}' nicht gefunden")
        try:
            return {"key": key, "value": json.loads(row["value"])}
        except Exception:
            return {"key": key, "value": row["value"]}

@app.put("/api/settings/{key}")
def upsert_setting(key: str, data: SettingUpsert):
    with db() as conn:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            (key, data.value)
        )
        return {"key": key, "value": data.value}


# ── Contracts ─────────────────────────────────────────────────────────────────

class ContractUpsert(BaseModel):
    gesellschaft:    Optional[str]   = ""
    police_nr:       Optional[str]   = ""
    ablaufdatum:     Optional[str]   = ""
    beitrag_alt:     Optional[float] = None
    absicherung_alt: Optional[str]   = ""
    beitrag_neu:     Optional[float] = None
    absicherung_neu: Optional[str]   = ""


@app.get("/api/customers/{customer_id}/contracts")
def get_contracts(customer_id: int):
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM contracts WHERE customer_id = %s ORDER BY id", (customer_id,)
        ).fetchall()
        result = {cat: [] for cat in CATEGORIES}
        for r in rows:
            cat = r["category"]
            if cat in result:
                result[cat].append(dict(r))
        return result


@app.post("/api/customers/{customer_id}/contracts/{category}", status_code=201)
def create_contract(customer_id: int, category: str, data: ContractUpsert):
    if category not in CATEGORIES:
        raise HTTPException(400, f"Unbekannte Kategorie: {category}")
    with db() as conn:
        cur = conn.execute("""
            INSERT INTO contracts (customer_id, category, gesellschaft, police_nr, ablaufdatum,
                                   beitrag_alt, absicherung_alt, beitrag_neu, absicherung_neu, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW()) RETURNING id
        """, (customer_id, category, data.gesellschaft, data.police_nr, data.ablaufdatum,
              data.beitrag_alt, data.absicherung_alt, data.beitrag_neu, data.absicherung_neu))
        new_id = cur.fetchone()["id"]
        return dict(conn.execute("SELECT * FROM contracts WHERE id = %s", (new_id,)).fetchone())


@app.put("/api/customers/{customer_id}/contracts/{contract_id}")
def update_contract(customer_id: int, contract_id: int, data: ContractUpsert):
    with db() as conn:
        conn.execute("""
            UPDATE contracts SET gesellschaft=%s, police_nr=%s, ablaufdatum=%s,
                beitrag_alt=%s, absicherung_alt=%s, beitrag_neu=%s, absicherung_neu=%s, updated_at=NOW()
            WHERE id=%s AND customer_id=%s
        """, (data.gesellschaft, data.police_nr, data.ablaufdatum,
              data.beitrag_alt, data.absicherung_alt, data.beitrag_neu, data.absicherung_neu,
              contract_id, customer_id))
        row = conn.execute("SELECT * FROM contracts WHERE id = %s", (contract_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Vertrag nicht gefunden")
        return dict(row)


@app.delete("/api/customers/{customer_id}/contracts/{contract_id}", status_code=204)
def delete_contract(customer_id: int, contract_id: int):
    with db() as conn:
        conn.execute("DELETE FROM contracts WHERE id=%s AND customer_id=%s", (contract_id, customer_id))


# ── Excel Export ──────────────────────────────────────────────────────────────

EXCEL_SECTIONS = {
    "Sach":     ["haftpflicht", "gebaeude", "hausrat", "rechtsschutz", "unfall", "tier"],
    "Kraft":    ["kfz"],
    "Vorsorge": ["einkommensschutz", "todesfall", "kindervorsorge", "alter", "vermoegen", "wohnwuensche", "immobilien"],
    "Gesundheit": ["gesundheit"],
}

EXCEL_LABELS = {
    "haftpflicht":      "Privathaftpflicht",
    "gebaeude":         "Wohngebäude",
    "hausrat":          "Hausrat",
    "rechtsschutz":     "Rechtsschutz",
    "unfall":           "Unfall",
    "tier":             "Tier",
    "kfz":              "KFZ",
    "einkommensschutz": "Arbeitskraft / BU",
    "todesfall":        "Risiko LV",
    "kindervorsorge":   "Kinderpolice",
    "alter":            "Altersvorsorge",
    "vermoegen":        "Vermögensaufbau",
    "wohnwuensche":     "Wohnen / Bausparen",
    "immobilien":       "Immobilien",
    "gesundheit":       "Gesundheit",
}


@app.get("/api/customers/{customer_id}/export/excel")
def export_excel(customer_id: int):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl nicht installiert")

    with db() as conn:
        kunde = conn.execute("SELECT * FROM customers WHERE id = %s", (customer_id,)).fetchone()
        if not kunde:
            raise HTTPException(404, "Kunde nicht gefunden")
        kunde = dict(kunde)
        contracts_rows = conn.execute(
            "SELECT * FROM contracts WHERE customer_id = %s", (customer_id,)
        ).fetchall()
        contracts = {r["category"]: dict(r) for r in contracts_rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Kundenübersicht"

    RED       = "C0392B"
    DARKRED   = "922B21"
    LIGHTGRAY = "F4F5F7"
    WHITE     = "FFFFFF"
    DARKTEXT  = "1A1A2E"

    def cell_style(cell, bold=False, size=11, color=DARKTEXT, bg=None,
                   align="left", wrap=False):
        cell.font = Font(name="Arial", bold=bold, size=size, color=color)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

    def money(val):
        return f"{val:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".") if val else "—"

    row = 1

    ws.merge_cells(f"A{row}:H{row}")
    name = f"{kunde.get('p1_vorname','')} {kunde.get('p1_nachname','')}".strip()
    if kunde.get("p2_vorname"):
        name += f" & {kunde.get('p2_vorname','')} {kunde.get('p2_nachname','')}".strip()
    c = ws.cell(row=row, column=1, value=name)
    cell_style(c, bold=True, size=16, color=RED)
    ws.row_dimensions[row].height = 30
    row += 1

    if kunde.get("p1_ort") or kunde.get("berater"):
        ws.merge_cells(f"A{row}:H{row}")
        sub = " · ".join(filter(None, [kunde.get("p1_ort"), f"Berater: {kunde.get('berater')}" if kunde.get("berater") else None]))
        c = ws.cell(row=row, column=1, value=sub)
        cell_style(c, size=10, color="888888")
        row += 1

    row += 1

    headers = ["Versicherung", "Ablauf", "Beitrag alt (€/mtl.)",
               "Beitrag neu (€/mtl.)", "Δ Beitrag", "Absicherung alt", "Absicherung neu", ""]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        cell_style(c, bold=True, size=10, color=WHITE, bg=DARKRED, align="center")
    ws.row_dimensions[row].height = 22
    row += 1

    total_alt = 0.0
    total_neu = 0.0

    for section, cats in EXCEL_SECTIONS.items():
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=section)
        cell_style(c, bold=True, size=11, color=WHITE, bg=RED)
        ws.row_dimensions[row].height = 20
        row += 1

        for cat in cats:
            ct = contracts.get(cat, {})
            b_alt = ct.get("beitrag_alt")
            b_neu = ct.get("beitrag_neu")
            delta = None
            if b_alt is not None and b_neu is not None:
                delta = b_neu - b_alt
            if b_alt:
                total_alt += b_alt
            if b_neu:
                total_neu += b_neu

            row_data = [
                EXCEL_LABELS.get(cat, cat),
                ct.get("ablaufdatum") or "",
                b_alt if b_alt is not None else "",
                b_neu if b_neu is not None else "",
                delta if delta is not None else "",
                ct.get("absicherung_alt") or "",
                ct.get("absicherung_neu") or "",
                "",
            ]
            bg = LIGHTGRAY if row % 2 == 0 else WHITE
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                align = "right" if col in (2, 3, 4, 5) else "left"
                cell_style(c, size=10, bg=bg, align=align, wrap=(col in (6, 7)))
                if col in (3, 4, 5) and isinstance(val, (int, float)):
                    c.number_format = '#,##0.00 "€"'
            ws.row_dimensions[row].height = 28
            row += 1

    row += 1

    for col, val in enumerate(["Gesamt", "", total_alt or "", total_neu or "",
                                (total_neu - total_alt) if (total_alt and total_neu) else "", "", "", ""], 1):
        c = ws.cell(row=row, column=col, value=val)
        cell_style(c, bold=True, size=11, bg="F0F0F0", align="right" if col > 1 else "left")
        if col in (3, 4, 5) and isinstance(val, (int, float)):
            c.number_format = '#,##0.00 "€"'
    ws.row_dimensions[row].height = 22
    row += 1

    if total_alt and total_neu:
        ersparnis = total_alt - total_neu
        ws.merge_cells(f"A{row}:E{row}")
        label = f"Ihre Ersparnis: {money(ersparnis)} mtl." if ersparnis > 0 else f"Mehrkosten: {money(-ersparnis)} mtl."
        c = ws.cell(row=row, column=1, value=label)
        cell_style(c, bold=True, size=12,
                   color=WHITE if ersparnis > 0 else DARKTEXT,
                   bg="1A7A4A" if ersparnis > 0 else "E8B800",
                   align="center")
        ws.row_dimensions[row].height = 26
        row += 2

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="Ihre Vorteile")
    cell_style(c, bold=True, size=12, color=WHITE, bg=RED)
    ws.row_dimensions[row].height = 22
    row += 1

    for v in [
        "Alle Versicherungen aus einer Hand",
        "Ein persönlicher Ansprechpartner",
        "Alle Versicherungen sind aufeinander abgestimmt",
        "Schadenexpertin im Team",
    ]:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=f"  ✓  {v}")
        cell_style(c, size=10, bg=LIGHTGRAY)
        ws.row_dimensions[row].height = 18
        row += 1

    for i, w in enumerate([28, 14, 20, 20, 14, 40, 40, 4], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today    = datetime.now().strftime("%Y-%m-%d")
    fullname = f"{kunde.get('p1_vorname','')} {kunde.get('p1_nachname','Kunde')}".strip()
    filename = f"{today} Analyse {fullname}.xlsx"
    encoded  = filename.encode('utf-8').decode('latin-1', errors='replace')
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"inline; filename=\"{encoded}\"; filename*=UTF-8''{filename.replace(' ', '%20')}"}
    )


# ── Tasks ─────────────────────────────────────────────────────────────────────

class TaskCreate(BaseModel):
    type:        str = "manual"
    title:       str
    description: str = ""
    customer_id: Optional[int] = None
    meta:        Optional[dict] = {}


@app.get("/api/tasks")
def list_tasks(status: str = "open"):
    with db() as conn:
        rows = conn.execute("""
            SELECT t.*, c.p1_vorname, c.p1_nachname
            FROM tasks t
            LEFT JOIN customers c ON c.id = t.customer_id
            WHERE t.status = %s
            ORDER BY t.created_at DESC
        """, (status,)).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            d["meta"] = json.loads(d.get("meta") or "{}")
            result.append(d)
        return result


@app.get("/api/tasks/count")
def count_tasks():
    with db() as conn:
        row = conn.execute("SELECT COUNT(*) as n FROM tasks WHERE status = 'open'").fetchone()
        return {"count": row["n"]}


@app.post("/api/tasks", status_code=201)
def create_task(data: TaskCreate):
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO tasks (type, title, description, customer_id, meta) VALUES (%s,%s,%s,%s,%s) RETURNING id",
            (data.type, data.title, data.description, data.customer_id, json.dumps(data.meta or {}))
        )
        task_id = cur.fetchone()["id"]
        row = conn.execute("SELECT * FROM tasks WHERE id = %s", (task_id,)).fetchone()
        d = dict(row)
        d["meta"] = json.loads(d.get("meta") or "{}")
        return d


@app.put("/api/tasks/{task_id}/done", status_code=200)
def mark_task_done(task_id: int):
    with db() as conn:
        conn.execute("UPDATE tasks SET status = 'done' WHERE id = %s", (task_id,))
        return {"ok": True}


@app.delete("/api/tasks/{task_id}", status_code=204)
def delete_task(task_id: int):
    with db() as conn:
        conn.execute("DELETE FROM tasks WHERE id = %s", (task_id,))


# ── Lead Import ───────────────────────────────────────────────────────────────

LEAD_EMAIL_PATTERN = re.compile(
    r"Lead-ID:\s*(?P<lead_id>\S+).*?"
    r"Kundenname:\s*(?P<name>.+?)\r?\n.*?"
    r"Telefon:\s*(?P<telefon>.+?)\r?\n.*?"
    r"E-Mail:\s*(?P<email>.+?)\r?\n.*?"
    r"Angefragtes Produkt:\s*(?P<produkt>.+?)\r?\n",
    re.DOTALL
)

DEFAULT_LEAD_TEMPLATE = """Hallo {{vorname}},

vielen Dank für Ihre Anfrage zur {{produkt}}! Ich freue mich über Ihr Interesse.

Ich werde mich in Kürze persönlich bei Ihnen melden. Wenn Sie möchten, können Sie auch direkt einen Wunschtermin buchen:
{{buchungslink}}

Mit freundlichen Grüßen
{{berater}}
Finanz-Team Schmidt GmbH"""


class LeadImport(BaseModel):
    raw_email: str


def _send_resend_email(to_email: str, from_label: str, from_email: str, subject: str, body: str) -> bool:
    if not RESEND_API_KEY:
        return False
    try:
        import resend
        resend.api_key = RESEND_API_KEY
        resend.Emails.send({
            "from": f"{from_label} <{from_email}>",
            "to": [to_email],
            "subject": subject,
            "text": body,
        })
        return True
    except Exception:
        return False


@app.post("/api/leads/import", status_code=201)
def import_lead(data: LeadImport):
    m = LEAD_EMAIL_PATTERN.search(data.raw_email)
    if not m:
        raise HTTPException(400, "E-Mail konnte nicht geparst werden")

    lead_id  = m.group("lead_id").strip()
    name     = m.group("name").strip()
    telefon  = m.group("telefon").strip()
    email    = m.group("email").strip()
    produkt  = m.group("produkt").strip()

    parts   = name.split()
    vorname  = parts[0] if parts else ""
    nachname = " ".join(parts[1:]) if len(parts) > 1 else ""

    with db() as conn:
        mapping_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'lead_produkt_mapping'"
        ).fetchone()
        mapping = json.loads(mapping_row["value"]) if mapping_row else []
        lookup  = {e.get("produkt", "").lower(): e.get("kategorie") for e in mapping}
        kategorie = lookup.get(produkt.lower())

        anfrage_kats = [kategorie] if kategorie else []
        cur = conn.execute("""
            INSERT INTO customers (p1_vorname, p1_nachname, p1_email, p1_handy,
                                   kontaktquelle, anfrage_kategorien, notizen)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (
            vorname, nachname, email, telefon,
            json.dumps(["Lead"]),
            json.dumps(anfrage_kats),
            f"Lead-ID: {lead_id} | Produkt: {produkt}"
        ))
        customer_id = cur.fetchone()["id"]

        if kategorie and kategorie in CATEGORIES:
            conn.execute("""
                INSERT INTO customer_categories (customer_id, category, status)
                VALUES (%s, %s, 'offen')
                ON CONFLICT (customer_id, category) DO UPDATE SET status='offen'
            """, (customer_id, kategorie))

        today = datetime.now().strftime("%Y-%m-%d")
        conn.execute(
            "INSERT INTO appointments (customer_id, datum, status, notizen) VALUES (%s,%s,%s,%s)",
            (customer_id, today, "Neuer Lead", f"Lead eingegangen: {produkt} (Lead-ID: {lead_id})")
        )

        if not kategorie:
            conn.execute("""
                INSERT INTO tasks (type, title, description, customer_id, meta)
                VALUES ('kategorie_zuordnen', %s, %s, %s, %s)
            """, (
                f"Kategorie zuordnen: {produkt}",
                f"Produkt '{produkt}' konnte keiner Kategorie zugeordnet werden. Bitte Mapping in Einstellungen ergänzen.",
                customer_id,
                json.dumps({"produkt": produkt, "lead_id": lead_id})
            ))

        email_sent = False
        template_row = conn.execute("SELECT value FROM settings WHERE key = 'lead_antwort_template'").fetchone()
        link_row     = conn.execute("SELECT value FROM settings WHERE key = 'lead_buchungslink'").fetchone()
        berater_row  = conn.execute("SELECT value FROM settings WHERE key = 'smtp_from_name'").fetchone()
        from_row     = conn.execute("SELECT value FROM settings WHERE key = 'smtp_config'").fetchone()

        if email:
            tmpl         = json.loads(template_row["value"]) if template_row else DEFAULT_LEAD_TEMPLATE
            link         = json.loads(link_row["value"]) if link_row else ""
            berater_name = json.loads(berater_row["value"]) if berater_row else ""
            from_cfg     = json.loads(from_row["value"]) if from_row else {}
            from_email   = from_cfg.get("user", "")

            body = tmpl \
                .replace("{{vorname}}", vorname) \
                .replace("{{produkt}}", produkt) \
                .replace("{{buchungslink}}", link) \
                .replace("{{berater}}", berater_name)

            email_sent = _send_resend_email(
                to_email=email,
                from_label=berater_name,
                from_email=from_email,
                subject=f"Ihre Anfrage zur {produkt}",
                body=body,
            )

            if email_sent:
                conn.execute(
                    "INSERT INTO appointments (customer_id, datum, status, notizen) VALUES (%s,%s,%s,%s)",
                    (customer_id, today, "Nachfassen", f"Automatische E-Mail gesendet an {email}")
                )
                conn.execute("""
                    INSERT INTO tasks (type, title, description, customer_id, meta)
                    VALUES ('termin_pruefen', %s, %s, %s, %s)
                """, (
                    f"Termin prüfen: {vorname} {nachname}",
                    f"Automatische Nachricht wurde gesendet. Bitte prüfen ob {vorname} einen Termin gebucht hat.",
                    customer_id,
                    json.dumps({"produkt": produkt, "email": email})
                ))

        return {
            "ok": True,
            "customer_id": customer_id,
            "kategorie": kategorie,
            "task_created": not kategorie,
            "email_sent": email_sent,
            "debug": {
                "parsed_produkt": produkt,
                "parsed_name": name,
                "mapping_keys": list(lookup.keys()),
                "match_found": kategorie is not None,
            }
        }


@app.post("/api/leads/{customer_id}/simulate-sent")
def simulate_sent(customer_id: int):
    with db() as conn:
        kunde = conn.execute(
            "SELECT p1_vorname, p1_nachname, p1_email FROM customers WHERE id = %s", (customer_id,)
        ).fetchone()
        if not kunde:
            raise HTTPException(404, "Kunde nicht gefunden")
        today = datetime.now().strftime("%Y-%m-%d")
        conn.execute(
            "INSERT INTO appointments (customer_id, datum, status, notizen) VALUES (%s,%s,%s,%s)",
            (customer_id, today, "Nachfassen", f"Automatische Nachricht gesendet an {kunde['p1_email'] or 'unbekannt'}")
        )
        conn.execute("""
            INSERT INTO tasks (type, title, description, customer_id, meta)
            VALUES ('termin_pruefen', %s, %s, %s, '{}')
        """, (
            f"Termin prüfen: {(kunde['p1_vorname'] or '')} {(kunde['p1_nachname'] or '')}".strip(),
            "Automatische Nachricht wurde gesendet. Bitte prüfen ob ein Termin gebucht wurde.",
            customer_id,
        ))
        return {"ok": True}


@app.get("/api/leads/unmapped-report")
def unmapped_report():
    with db() as conn:
        mapping_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'lead_produkt_mapping'"
        ).fetchone()
        mapping = json.loads(mapping_row["value"]) if mapping_row else []
        known_products = {e["produkt"].lower() for e in mapping}

        rows = conn.execute("""
            SELECT id, p1_vorname, p1_nachname, notizen, anfrage_kategorien, created_at
            FROM customers WHERE notizen LIKE '%Produkt:%'
        """).fetchall()

        result = []
        for row in rows:
            notizen = row["notizen"] or ""
            m = re.search(r"Lead-ID:\s*(\S+).*?Produkt:\s*(.+?)($|\s*\|)", notizen)
            if not m:
                continue
            lead_id  = m.group(1).strip()
            produkt  = m.group(2).strip()
            mapped   = produkt.lower() in known_products
            kats     = json.loads(row["anfrage_kategorien"] or "[]")
            result.append({
                "customer_id": row["id"],
                "name": f"{row['p1_vorname']} {row['p1_nachname']}".strip(),
                "produkt": produkt,
                "lead_id": lead_id,
                "mapped": mapped,
                "has_kategorie": len(kats) > 0,
                "created_at": str(row["created_at"]),
            })
        return result


@app.post("/api/leads/apply-mapping")
def apply_mapping():
    with db() as conn:
        mapping_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'lead_produkt_mapping'"
        ).fetchone()
        if not mapping_row:
            return {"applied": 0}

        mapping = json.loads(mapping_row["value"])
        lookup  = {e["produkt"].lower(): e["kategorie"] for e in mapping}
        applied = 0

        open_tasks = conn.execute(
            "SELECT * FROM tasks WHERE type = 'kategorie_zuordnen' AND status = 'open'"
        ).fetchall()

        for task in open_tasks:
            meta        = json.loads(task["meta"] or "{}")
            produkt_key = meta.get("produkt", "").lower()
            if produkt_key not in lookup:
                continue
            kategorie   = lookup[produkt_key]
            customer_id = task["customer_id"]
            if not customer_id or kategorie not in CATEGORIES:
                continue
            _apply_kategorie(conn, customer_id, kategorie)
            conn.execute("UPDATE tasks SET status = 'done' WHERE id = %s", (task["id"],))
            applied += 1

        all_leads = conn.execute(
            "SELECT id, notizen, anfrage_kategorien FROM customers WHERE notizen LIKE '%Produkt:%'"
        ).fetchall()
        for row in all_leads:
            notizen = row["notizen"] or ""
            mat = re.search(r"Produkt:\s*(.+?)($|\s*\|)", notizen)
            if not mat:
                continue
            produkt_key = mat.group(1).strip().lower()
            if produkt_key not in lookup:
                continue
            kategorie = lookup[produkt_key]
            if kategorie not in CATEGORIES:
                continue
            kats = json.loads(row["anfrage_kategorien"] or "[]")
            if kategorie not in kats:
                _apply_kategorie(conn, row["id"], kategorie)
                applied += 1

        return {"applied": applied}


def _apply_kategorie(conn, customer_id: int, kategorie: str):
    row = conn.execute("SELECT anfrage_kategorien FROM customers WHERE id = %s", (customer_id,)).fetchone()
    if row:
        kats = json.loads(row["anfrage_kategorien"] or "[]")
        if kategorie not in kats:
            kats.append(kategorie)
            conn.execute("UPDATE customers SET anfrage_kategorien = %s WHERE id = %s",
                         (json.dumps(kats), customer_id))
    conn.execute("""
        INSERT INTO customer_categories (customer_id, category, status)
        VALUES (%s, %s, 'offen')
        ON CONFLICT (customer_id, category) DO UPDATE SET status='offen'
    """, (customer_id, kategorie))
